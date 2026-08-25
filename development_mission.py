from pathlib import Path
from datetime import datetime
from collections import defaultdict
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# 이 파이썬 파일이 들어 있는 폴더를 기준 경로로 사용합니다.
# 따라서 VS Code 터미널이 다른 폴더에 있어도 파일 경로가 흔들리지 않습니다.
BASE_DIR = Path(__file__).resolve().parent
INPUT_FOLDER = BASE_DIR / "advanced_input"
OUTPUT_FOLDER = BASE_DIR / "advanced_output"


# 기관마다 다른 원본 컬럼을 공통 컬럼으로 바꾸기 위한 규칙입니다.
# 왼쪽은 원본 CSV의 컬럼명이고, 오른쪽은 프로그램에서 사용할 표준 컬럼명입니다.
COLUMN_RULES = {
    "카드사 A": {
        "이용일": "거래일자",
        "가맹점명": "거래처",
        "이용금액": "금액",
        "카드구분": "결제수단",
        "비고": "메모",
    },
    "은행": {
        "거래일시": "거래일자",
        "적요": "거래처",
        "출금액": "금액",
        "거래구분": "결제수단",
        "내용": "메모",
    },
    "카드사 B": {
        "승인일자": "거래일자",
        "이용가맹점": "거래처",
        "승인금액": "금액",
        "지급수단": "결제수단",
        "이용구분": "메모",
    },
}


# 표준화가 끝난 거래처명을 이용해 계정과목을 분류합니다.
# 이 부분은 앞의 기본 프로젝트에서 사용한 규칙과 같은 역할을 합니다.
ACCOUNT_RULES = {
    "스타벅스": "복리후생비",
    "배달의민족": "복리후생비",
    "쿠팡": "소모품비",
    "알파문구": "소모품비",
    "코레일": "여비교통비",
    "한국전력": "수도광열비",
    "사무실임대": "임차료",
    "KT": "통신비",
    "네이버클라우드": "통신비",
    "국세청": "세금과공과",
    "세무법인": "지급수수료",
}


# 최종 CSV와 Excel 시트에 기록할 표준 컬럼과 순서입니다.
STANDARD_COLUMNS = [
    "원본파일",
    "원본형식",
    "거래일자",
    "거래처",
    "금액",
    "결제수단",
    "메모",
    "계정과목",
    "처리상태",
]


def detect_file_format(fieldnames):
    """CSV의 컬럼 조합을 검사하여 어느 기관 형식인지 판별합니다."""

    # DictReader가 읽은 컬럼명을 집합으로 바꾸면 포함 여부를 쉽게 검사할 수 있습니다.
    file_columns = set(fieldnames or [])

    # 등록된 기관별 규칙을 하나씩 확인합니다.
    for format_name, column_map in COLUMN_RULES.items():
        required_columns = set(column_map.keys())

        # 현재 파일이 해당 기관의 필수 컬럼을 모두 가지고 있으면 그 규칙을 반환합니다.
        # 원본 파일에 추가 컬럼이 더 있어도 필수 컬럼만 있으면 처리할 수 있습니다.
        if required_columns.issubset(file_columns):
            return format_name, column_map

    # 어떤 규칙과도 맞지 않으면 억지로 합치지 않고 오류를 발생시킵니다.
    raise ValueError(
        "등록되지 않은 컬럼 구조입니다: "
        + ", ".join(fieldnames or [])
    )


def normalize_date(value):
    """여러 형태의 날짜를 YYYY-MM-DD 형식으로 통일합니다."""

    value = value.strip()

    # 실습파일에서 사용한 날짜 형식을 차례대로 대입해 봅니다.
    date_formats = [
        "%Y-%m-%d",
        "%Y.%m.%d",
        "%Y/%m/%d",
        "%Y%m%d",
        "%Y%m%d %H:%M",
    ]

    for date_format in date_formats:
        try:
            parsed_date = datetime.strptime(value, date_format)
            return parsed_date.strftime("%Y-%m-%d")
        except ValueError:
            # 현재 형식과 맞지 않으면 다음 날짜 형식으로 다시 시도합니다.
            continue

    raise ValueError(f"지원하지 않는 날짜 형식: {value}")


def normalize_amount(value):
    """금액의 쉼표·원·공백을 제거하고 계산 가능한 정수로 바꿉니다."""

    cleaned_value = (
        value.strip()
        .replace(",", "")
        .replace("원", "")
        .replace(" ", "")
    )

    try:
        return int(cleaned_value)
    except ValueError as error:
        raise ValueError(f"금액을 숫자로 바꿀 수 없음: {value}") from error


def classify_account(merchant):
    """거래처 키워드를 이용하여 계정과목을 반환합니다."""

    for keyword, account in ACCOUNT_RULES.items():
        if keyword in merchant:
            return account

    return "확인 필요"


def standardize_row(row, file_name, format_name, column_map):
    """서로 다른 원본 행을 하나의 표준 거래내역으로 변환합니다."""

    standardized = {}

    # 원본 컬럼의 값을 가져와 대응하는 표준 컬럼명으로 저장합니다.
    for original_column, standard_column in column_map.items():
        standardized[standard_column] = row.get(original_column, "").strip()

    # 필수값이 비어 있으면 잘못된 자료가 섞이지 않도록 중단합니다.
    if not standardized["거래일자"]:
        raise ValueError("거래일자 누락")
    if not standardized["거래처"]:
        raise ValueError("거래처 누락")
    if not standardized["금액"]:
        raise ValueError("금액 누락")

    # 컬럼 이름뿐 아니라 날짜와 금액의 값 형식도 통일합니다.
    standardized["거래일자"] = normalize_date(standardized["거래일자"])
    standardized["금액"] = normalize_amount(standardized["금액"])

    # 표준화된 거래처명을 이용해 기존 분류 규칙을 적용합니다.
    account = classify_account(standardized["거래처"])
    status = "정상" if account != "확인 필요" else "계정과목 확인"

    # 최종 결과는 모든 원본 파일에서 동일한 컬럼과 순서를 갖습니다.
    return {
        "원본파일": file_name,
        "원본형식": format_name,
        "거래일자": standardized["거래일자"],
        "거래처": standardized["거래처"],
        "금액": standardized["금액"],
        "결제수단": standardized["결제수단"],
        "메모": standardized["메모"],
        "계정과목": account,
        "처리상태": status,
    }


def load_and_standardize_files():
    """advanced_input의 모든 CSV를 읽어 표준 거래내역으로 통합합니다."""

    csv_files = sorted(INPUT_FOLDER.glob("*.csv"))

    if not csv_files:
        print(f"CSV 파일이 없습니다: {INPUT_FOLDER}")
        return []

    all_rows = []

    for csv_file in csv_files:
        with open(csv_file, "r", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)

            # 파일명이 아니라 실제 컬럼 조합으로 기관 형식을 판별합니다.
            format_name, column_map = detect_file_format(reader.fieldnames)
            print(f"{csv_file.name} → {format_name} 형식으로 판별")

            for line_number, row in enumerate(reader, start=2):
                try:
                    standardized_row = standardize_row(
                        row,
                        csv_file.name,
                        format_name,
                        column_map,
                    )
                    all_rows.append(standardized_row)
                except ValueError as error:
                    # 오류가 난 행과 이유를 화면에 알려 주고 다음 행을 계속 처리합니다.
                    print(f"  {line_number}행 제외: {error}")

    # 파일 종류와 관계없이 최종 결과를 거래일자순으로 정렬합니다.
    all_rows.sort(key=lambda row: (row["거래일자"], row["원본파일"]))
    return all_rows


def save_standard_csv(rows):
    """표준화한 전체 거래를 하나의 CSV 파일로 저장합니다."""

    csv_path = OUTPUT_FOLDER / "standardized_transactions.csv"

    with open(
        csv_path,
        "w",
        newline="",
        encoding="utf-8-sig",
    ) as file:
        writer = csv.DictWriter(file, fieldnames=STANDARD_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)

    return csv_path


def style_sheet(sheet, header_color):
    """Excel 시트의 제목과 열 너비에 공통 서식을 적용합니다."""

    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(
            width,
            30,
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def create_excel_report(rows):
    """표준 거래·계정과목별 합계·확인 필요 시트가 있는 보고서를 만듭니다."""

    workbook = Workbook()

    # 첫 번째 시트에는 표준화된 전체 거래를 기록합니다.
    transaction_sheet = workbook.active
    transaction_sheet.title = "표준 거래내역"
    transaction_sheet.append(STANDARD_COLUMNS)

    for row in rows:
        transaction_sheet.append([row[column] for column in STANDARD_COLUMNS])

    # 두 번째 시트에는 계정과목별 거래 건수와 합계를 기록합니다.
    totals = defaultdict(int)
    counts = defaultdict(int)

    for row in rows:
        totals[row["계정과목"]] += row["금액"]
        counts[row["계정과목"]] += 1

    summary_sheet = workbook.create_sheet("계정과목별 합계")
    summary_sheet.append(["계정과목", "거래 건수", "합계 금액"])

    for account in sorted(totals):
        summary_sheet.append([account, counts[account], totals[account]])

    summary_sheet.append(
        ["전체 합계", sum(counts.values()), sum(totals.values())]
    )

    # 세 번째 시트에는 사람이 계정과목을 확인해야 할 거래만 기록합니다.
    review_sheet = workbook.create_sheet("확인 필요 거래")
    review_sheet.append(STANDARD_COLUMNS)

    for row in rows:
        if row["처리상태"] != "정상":
            review_sheet.append([row[column] for column in STANDARD_COLUMNS])

    style_sheet(transaction_sheet, "1D4ED8")
    style_sheet(summary_sheet, "6D28D9")
    style_sheet(review_sheet, "EA580C")

    # 금액 열에는 천 단위 쉼표 서식을 적용합니다.
    for cell in transaction_sheet["E"][1:]:
        cell.number_format = "#,##0"
    for cell in summary_sheet["C"][1:]:
        cell.number_format = "#,##0"

    report_path = OUTPUT_FOLDER / "different_columns_report.xlsx"
    workbook.save(report_path)
    return report_path


def main():
    """컬럼 판별부터 표준화·통합·보고서 생성까지 한 번에 실행합니다."""

    OUTPUT_FOLDER.mkdir(exist_ok=True)
    rows = load_and_standardize_files()

    if not rows:
        return

    csv_path = save_standard_csv(rows)
    report_path = create_excel_report(rows)
    total_amount = sum(row["금액"] for row in rows)

    print()
    print(f"표준화 완료: {len(rows)}건")
    print(f"전체 합계: {total_amount:,}원")
    print(f"CSV 저장: {csv_path}")
    print(f"Excel 저장: {report_path}")


if __name__ == "__main__":
    main()
